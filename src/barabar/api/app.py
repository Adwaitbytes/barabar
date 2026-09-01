"""FastAPI service: ingest, runs, exceptions, proof trees, exports, chat gateway,
webhooks. Thin by design — every decision lives in ``barabar.core``."""

from __future__ import annotations

import csv
import io
import json
import os
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any

from fastapi import Depends, FastAPI, File, Form, Header, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import PlainTextResponse, Response
from pydantic import BaseModel, Field

from barabar.adapters.bank_csv import read_bank_csv, read_bank_rows
from barabar.adapters.ledger_csv import read_ledger_csv
from barabar.adapters.razorpay_json import FILES
from barabar.api.store import Store
from barabar.api.webhooks import event_identity, verify_signature
from barabar.core.calendar import IST
from barabar.core.config import MatchConfig
from barabar.core.exceptions import EXCEPTION_SPECS
from barabar.core.models import ExceptionStatus, Month, SettlementStatus
from barabar.core.money import format_inr
from barabar.core.result import ReconResult
from barabar.evals.datasets import read_dataset
from barabar.exports.journal import journal_csv, vouchers_for_run
from barabar.exports.memo import controller_memo, memo_facts
from barabar.exports.tally_xml import tally_xml
from barabar.generator.engine import generate
from barabar.generator.profiles import MerchantProfile

APP_NAME = "Barabar"

app = FastAPI(title=f"{APP_NAME} API", version="0.1.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

_store: Store | None = None
_cfg = MatchConfig()


def store() -> Store:
    global _store
    if _store is None:
        _store = Store()
        _store.resume_incomplete(_cfg)
    return _store


def cfg() -> MatchConfig:
    return _cfg


# --- schemas --------------------------------------------------------------------


class CreateRun(BaseModel):
    source: str = Field(pattern="^(synthetic|dataset)$")
    n_orders: int = 600
    seed: int = 42
    profile: str = MerchantProfile.D2C_FASHION.value
    dataset_path: str | None = None
    name: str | None = None


class AskBody(BaseModel):
    question: str = Field(min_length=3, max_length=2000)


class ResolveBody(BaseModel):
    status: str = Field(pattern="^(resolved|accepted|open|investigating)$")
    note: str | None = None
    actor: str = "user:demo"


# --- helpers -----------------------------------------------------------------------


def _run_payload(s: Store, run_id: str) -> dict[str, Any]:
    meta = s.run_meta(run_id)
    return {
        "run_id": run_id,
        "name": meta["name"],
        "as_of": meta["as_of"],
        "stage": meta["stage"],
        "inputs_hash": meta["inputs_hash"],
        "config_hash": meta["config_hash"],
        "code_version": meta["code_version"],
        "outputs_hash": meta["outputs_hash"],
        "started_at": meta["started_at"].isoformat() if meta["started_at"] else None,
        "finished_at": meta["finished_at"].isoformat() if meta["finished_at"] else None,
        "metrics": meta["metrics"],
        "source": meta["source"],
    }


def _settlement_status(result: ReconResult, sid: str) -> str:
    sref = f"settlement:{sid}"
    if any(
        link.to_entity == sref and link.from_entity.startswith("bank:") for link in result.links
    ):
        for e in result.exceptions:
            if sref in e.entities and e.type.value in ("PARTIAL_SETTLEMENT", "MULTI_UTR_SPLIT"):
                return "partial" if e.type.value == "PARTIAL_SETTLEMENT" else "split"
        return "matched"
    for e in result.exceptions:
        if e.entities and e.entities[0] == sref:
            return {
                "TIMING_BANK_LAG": "pending",
                "MISSING_BANK_CREDIT": "missing",
                "NARRATION_TRUNCATED_UTR": "proposed",
                "SETTLEMENT_FAILED_RETURNED": "failed",
                "DUPLICATE_BANK_CREDIT": "duplicate",
            }.get(e.type.value, "open")
    return "unmatched"


def _exception_payload(e: Any) -> dict[str, Any]:
    d = e.model_dump(mode="json")
    d["amount_display"] = format_inr(e.amount)
    d["spec"] = {
        "meaning": EXCEPTION_SPECS[e.type].meaning,
        "auto_resolvable": EXCEPTION_SPECS[e.type].auto_resolvable,
    }
    return d


def _close_pack(month: Month, result: ReconResult) -> dict[str, Any]:
    m = result.metrics
    settlements = []
    for s in sorted(
        month.settlements, key=lambda s: (s.settled_at or s.created_at, s.settlement_id)
    ):
        lines = [
            ln
            for ln in month.recon_lines
            if ln.settlement_id == s.settlement_id and ln.settled and not ln.on_hold
        ]
        settlements.append(
            {
                "settlement_id": s.settlement_id,
                "amount": s.amount,
                "amount_display": format_inr(s.amount),
                "utr": s.utr,
                "status": s.status.value,
                "type": s.type.value,
                "mode": s.mode.value,
                "settled_at": s.settled_at.isoformat() if s.settled_at else None,
                "settled_on": (s.settled_at or s.created_at).astimezone(IST).date().isoformat(),
                "lines": len(lines),
                "gross": sum(ln.amount for ln in lines if ln.type.value == "payment"),
                "fee": sum(ln.fee for ln in lines),
                "tax": sum(ln.tax for ln in lines),
                "match_status": _settlement_status(result, s.settlement_id),
            }
        )
    expected: dict[str, int] = defaultdict(int)
    for s in month.settlements:
        if s.status == SettlementStatus.PROCESSED:
            expected[(s.settled_at or s.created_at).astimezone(IST).date().isoformat()] += s.amount
    actual: dict[str, int] = defaultdict(int)
    for t in month.bank_txns:
        if t.credit and t.narration and t.narration.razorpay_like:
            actual[t.value_date.isoformat()] += t.credit
    days = sorted(set(expected) | set(actual))
    calendar = [
        {
            "date": d,
            "expected": expected.get(d, 0),
            "actual": actual.get(d, 0),
            "delta": actual.get(d, 0) - expected.get(d, 0),
        }
        for d in days
    ]
    open_exc = [e for e in result.exceptions if e.status == ExceptionStatus.OPEN]
    by_type: dict[str, dict[str, int]] = defaultdict(lambda: {"count": 0, "amount": 0})
    for e in open_exc:
        by_type[e.type.value]["count"] += 1
        by_type[e.type.value]["amount"] += e.amount
    return {
        "headline": {
            "gross_captured": m["gross_captured_paise"],
            "explained": m["explained_paise"],
            "unexplained": m["unexplained_paise"],
            "rupees_explained_pct": m["rupees_explained_pct"],
            "ledger_open": m.get("ledger_open_paise", 0),
            "gross_captured_display": format_inr(int(m["gross_captured_paise"])),
            "explained_display": format_inr(int(m["explained_paise"])),
            "unexplained_display": format_inr(int(m["unexplained_paise"])),
        },
        "metrics": m,
        "settlements": settlements,
        "calendar": calendar,
        "exceptions_by_type": dict(by_type),
        "exceptions_open": len(open_exc),
        "exceptions_total": len(result.exceptions),
        "facts": memo_facts(month, result),
    }


# --- routes ------------------------------------------------------------------------------


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok", "app": APP_NAME}


@app.get("/runs")
def list_runs(s: Store = Depends(store)) -> list[dict[str, Any]]:
    return [_run_payload(s, r["run_id"]) for r in s.list_runs()]


@app.post("/runs", status_code=201)
def create_run(
    body: CreateRun, s: Store = Depends(store), c: MatchConfig = Depends(cfg)
) -> dict[str, Any]:
    if body.source == "synthetic":
        gen = generate(
            seed=body.seed, n_orders=body.n_orders, profile=MerchantProfile(body.profile)
        )
        month, source = (
            gen.month,
            {
                "kind": "synthetic",
                "seed": body.seed,
                "n_orders": body.n_orders,
                "profile": body.profile,
            },
        )
    else:
        if not body.dataset_path:
            raise HTTPException(400, "dataset_path required")
        month, _ = read_dataset(Path(body.dataset_path))
        source = {"kind": "dataset", "path": body.dataset_path}
    run = s.create_run(month, c, source=source, name=body.name)
    s.reconcile_run(run.run_id, c)
    return _run_payload(s, run.run_id)


@app.post("/runs/upload", status_code=201)
async def create_run_upload(
    as_of: str = Form(...),
    name: str | None = Form(None),
    razorpay: list[UploadFile] = File(default=[]),
    bank: UploadFile | None = File(None),
    ledger: UploadFile | None = File(None),
    s: Store = Depends(store),
    c: MatchConfig = Depends(cfg),
) -> dict[str, Any]:
    parts: dict[str, tuple[Any, ...]] = {k: () for k in FILES}
    for up in razorpay:
        raw = json.loads((await up.read()).decode("utf-8"))
        items = raw["items"] if isinstance(raw, dict) and "items" in raw else raw
        matched = False
        for attr, (fname, _, from_api) in FILES.items():
            if up.filename and (fname == up.filename or attr in up.filename):
                parts[attr] = tuple(from_api(i) for i in items)
                matched = True
                break
        if not matched:
            raise HTTPException(
                400,
                f"cannot tell which Razorpay entity {up.filename} holds; name it like {list(f[0] for f in FILES.values())}",
            )
    bank_txns: list[Any] = []
    if bank is not None:
        data = await bank.read()
        if bank.filename and bank.filename.lower().endswith((".xlsx", ".xlsm")):
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                raise HTTPException(400, "workbook has no active sheet")
            rows = [
                ["" if cell is None else str(cell) for cell in r]
                for r in ws.iter_rows(values_only=True)
            ]
            bank_txns = read_bank_rows(rows, source_file=bank.filename)
        else:
            try:
                bank_txns = read_bank_csv(
                    data.decode("utf-8-sig"), source_file=bank.filename or "bank.csv"
                )
            except ValueError as exc:
                raise HTTPException(400, str(exc)) from exc
    ledger_rows: list[Any] = []
    if ledger is not None:
        try:
            ledger_rows = read_ledger_csv((await ledger.read()).decode("utf-8-sig"))
        except ValueError as exc:
            raise HTTPException(400, str(exc)) from exc
    month = Month(
        as_of=date.fromisoformat(as_of),
        bank_txns=tuple(bank_txns),
        ledger=tuple(ledger_rows),
        **parts,
    )
    run = s.create_run(
        month,
        c,
        source={
            "kind": "upload",
            "files": [u.filename for u in razorpay] + [f.filename for f in (bank, ledger) if f],
        },
        name=name,
    )
    s.reconcile_run(run.run_id, c)
    return _run_payload(s, run.run_id)


@app.get("/runs/{run_id}")
def get_run(run_id: str, s: Store = Depends(store)) -> dict[str, Any]:
    try:
        return _run_payload(s, run_id)
    except KeyError as exc:
        raise HTTPException(404, "run not found") from exc


@app.delete("/runs/{run_id}", status_code=204)
def delete_run(run_id: str, s: Store = Depends(store)) -> Response:
    s.delete_run(run_id)
    return Response(status_code=204)


@app.post("/runs/{run_id}/rerun", status_code=201)
def rerun(run_id: str, s: Store = Depends(store), c: MatchConfig = Depends(cfg)) -> dict[str, Any]:
    old = s.get_run(run_id)
    month = s.get_month(old.inputs_hash)
    new = s.create_run(month, c, source={"kind": "rerun", "of": run_id}, name=f"re-run of {run_id}")
    s.reconcile_run(new.run_id, c)
    payload = _run_payload(s, new.run_id)
    payload["identical_outputs"] = payload["outputs_hash"] == old.outputs_hash
    payload["diff"] = _diff(s.get_result(run_id), s.get_result(new.run_id))
    return payload


def _diff(a: ReconResult, b: ReconResult) -> dict[str, Any]:
    ka = {e.exc_id: e for e in a.exceptions}
    kb = {e.exc_id: e for e in b.exceptions}
    return {
        "closed": [_exception_payload(ka[k]) for k in ka.keys() - kb.keys()],
        "opened": [_exception_payload(kb[k]) for k in kb.keys() - ka.keys()],
        "unchanged": len(ka.keys() & kb.keys()),
    }


@app.get("/runs/{run_id}/diff/{other_id}")
def diff_runs(run_id: str, other_id: str, s: Store = Depends(store)) -> dict[str, Any]:
    return _diff(s.get_result(run_id), s.get_result(other_id))


@app.get("/runs/{run_id}/close-pack")
def close_pack(run_id: str, s: Store = Depends(store)) -> dict[str, Any]:
    try:
        run = s.get_run(run_id)
    except KeyError as exc:
        raise HTTPException(404, "run not found") from exc
    return {
        "run": _run_payload(s, run_id),
        **_close_pack(s.get_month(run.inputs_hash), s.get_result(run_id)),
    }


@app.get("/runs/{run_id}/exceptions")
def list_exceptions(
    run_id: str, status: str | None = None, type: str | None = None, s: Store = Depends(store)
) -> list[dict[str, Any]]:
    result = s.get_result(run_id)
    items = [
        e
        for e in result.exceptions
        if (status is None or e.status.value == status) and (type is None or e.type.value == type)
    ]
    return [_exception_payload(e) for e in sorted(items, key=lambda e: (-e.amount, e.type.value))]


@app.get("/runs/{run_id}/exceptions/{exc_id}")
def get_exception(run_id: str, exc_id: str, s: Store = Depends(store)) -> dict[str, Any]:
    result = s.get_result(run_id)
    item = next((e for e in result.exceptions if e.exc_id == exc_id), None)
    if item is None:
        raise HTTPException(404, "exception not found")
    return _exception_payload(item)


@app.post("/runs/{run_id}/exceptions/{exc_id}/resolve")
def resolve_exception(
    run_id: str, exc_id: str, body: ResolveBody, s: Store = Depends(store)
) -> dict[str, Any]:
    try:
        item = s.set_exception_status(
            run_id, exc_id, ExceptionStatus(body.status), actor=body.actor, note=body.note
        )
    except KeyError as exc:
        raise HTTPException(404, "exception not found") from exc
    return _exception_payload(item)


@app.post("/runs/{run_id}/exceptions/{exc_id}/investigate")
def investigate_exception(
    run_id: str, exc_id: str, s: Store = Depends(store), c: MatchConfig = Depends(cfg)
) -> dict[str, Any]:
    from barabar.agent.investigator import InvestigatorUnavailableError, investigate

    run = s.get_run(run_id)
    result = s.get_result(run_id)
    if not any(e.exc_id == exc_id for e in result.exceptions):
        raise HTTPException(404, "exception not found")
    try:
        card, cached, model = investigate(
            s.get_month(run.inputs_hash), result, c, exc_id, client=_agent_client()
        )
    except InvestigatorUnavailableError as exc:
        raise HTTPException(503, str(exc)) from exc
    payload = card.as_dict()
    s.put_hypothesis(run_id, exc_id, payload, model)
    s.set_exception_status(
        run_id,
        exc_id,
        ExceptionStatus.INVESTIGATING,
        actor="agent",
        note=f"hypothesis {card.type_proposed.value} @ {card.confidence:.2f}; {len(card.evidence)} evidence items",
    )
    return {"exc_id": exc_id, "hypothesis": payload, "model": model, "cached": cached}


@app.get("/runs/{run_id}/exceptions/{exc_id}/hypothesis")
def get_hypothesis(run_id: str, exc_id: str, s: Store = Depends(store)) -> dict[str, Any]:
    h = s.get_hypothesis(run_id, exc_id)
    if h is None:
        raise HTTPException(404, "no hypothesis yet")
    return {"exc_id": exc_id, **h, "cached": True}


@app.post("/runs/{run_id}/ask")
def ask_the_books(
    run_id: str, body: AskBody, s: Store = Depends(store), c: MatchConfig = Depends(cfg)
) -> dict[str, Any]:
    from barabar.agent.investigator import InvestigatorUnavailableError, ask

    run = s.get_run(run_id)
    try:
        answer, model = ask(
            s.get_month(run.inputs_hash),
            s.get_result(run_id),
            c,
            body.question,
            client=_agent_client(),
        )
    except InvestigatorUnavailableError as exc:
        raise HTTPException(503, str(exc)) from exc
    s.append_audit(
        run_id,
        actor="agent",
        action="ask.answered",
        target=body.question[:120],
        rule_or_evidence=f"{len(answer.citations)} tool calls; guard checked {answer.guard.numbers_checked}, blocked={answer.guard.blocked}",
    )
    return {**answer.as_dict(), "model": model}


_agent_client_override: Any = None


def _agent_client() -> Any:
    return _agent_client_override


@app.get("/runs/{run_id}/proof/{settlement_id}")
def proof_tree(run_id: str, settlement_id: str, s: Store = Depends(store)) -> dict[str, Any]:
    result = s.get_result(run_id)
    node = result.proof_trees.get(settlement_id)
    if node is None:
        raise HTTPException(404, "no proof tree for that settlement")
    return node.model_dump(mode="json")


@app.get("/runs/{run_id}/proof-by-bank/{bank_txn_id}")
def proof_by_bank(run_id: str, bank_txn_id: str, s: Store = Depends(store)) -> dict[str, Any]:
    result = s.get_result(run_id)
    for link in result.links:
        if link.from_entity == f"bank:{bank_txn_id}" and link.to_entity.startswith("settlement:"):
            return result.proof_trees[link.to_entity.split(":", 1)[1]].model_dump(mode="json")
    raise HTTPException(404, "bank credit is not linked to a settlement")


@app.get("/runs/{run_id}/links")
def links(
    run_id: str, entity: str | None = None, s: Store = Depends(store)
) -> list[dict[str, Any]]:
    result = s.get_result(run_id)
    return [
        link.model_dump(mode="json")
        for link in result.links
        if entity is None or entity in (link.from_entity, link.to_entity)
    ]


@app.get("/runs/{run_id}/audit")
def audit_trail(run_id: str, s: Store = Depends(store)) -> dict[str, Any]:
    chain = s.audit_chain(run_id)
    return {
        "head": chain.head,
        "verified": chain.verify(),
        "events": [e.model_dump(mode="json") for e in chain.events],
    }


@app.get("/runs/{run_id}/month")
def month_entities(run_id: str, kind: str, s: Store = Depends(store)) -> list[dict[str, Any]]:
    run = s.get_run(run_id)
    month = s.get_month(run.inputs_hash)
    if kind not in (
        "payments",
        "refunds",
        "disputes",
        "adjustments",
        "settlements",
        "recon_lines",
        "bank_txns",
        "ledger",
    ):
        raise HTTPException(400, "unknown kind")
    return [x.model_dump(mode="json") for x in getattr(month, kind)]


# --- exports -----------------------------------------------------------------------------


@app.get("/runs/{run_id}/export/journal.csv")
def export_journal_csv(run_id: str, s: Store = Depends(store)) -> Response:
    run = s.get_run(run_id)
    text = journal_csv(vouchers_for_run(s.get_month(run.inputs_hash), s.get_result(run_id)))
    return Response(
        text,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{run_id}-journal.csv"'},
    )


@app.get("/runs/{run_id}/export/tally.xml")
def export_tally(run_id: str, gst_split: str = "igst", s: Store = Depends(store)) -> Response:
    run = s.get_run(run_id)
    text = tally_xml(
        vouchers_for_run(
            s.get_month(run.inputs_hash),
            s.get_result(run_id),
            gst_split="cgst_sgst" if gst_split == "cgst_sgst" else "igst",
        )
    )
    return Response(
        text,
        media_type="application/xml",
        headers={"Content-Disposition": f'attachment; filename="{run_id}-tally.xml"'},
    )


@app.get("/runs/{run_id}/export/memo.md", response_class=PlainTextResponse)
def export_memo(run_id: str, s: Store = Depends(store)) -> str:
    run = s.get_run(run_id)
    return controller_memo(s.get_month(run.inputs_hash), s.get_result(run_id))


@app.get("/runs/{run_id}/export/exceptions.csv")
def export_exceptions(run_id: str, s: Store = Depends(store)) -> Response:
    result = s.get_result(run_id)
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\n")
    w.writerow(
        [
            "exc_id",
            "type",
            "status",
            "amount",
            "confidence",
            "entities",
            "reason",
            "suggested_action",
        ]
    )
    for e in result.exceptions:
        w.writerow(
            [
                e.exc_id,
                e.type.value,
                e.status.value,
                format_inr(e.amount, symbol=False),
                e.confidence,
                " ".join(e.entities),
                e.reason_text,
                e.suggested_action,
            ]
        )
    return Response(
        buf.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{run_id}-exceptions.csv"'},
    )


# --- webhooks -------------------------------------------------------------------------------


@app.post("/webhooks/razorpay")
async def razorpay_webhook(
    request: Request,
    x_razorpay_signature: str | None = Header(None),
    x_razorpay_event_id: str | None = Header(None),
    s: Store = Depends(store),
) -> dict[str, Any]:
    body = await request.body()
    secret = os.environ.get("RAZORPAY_WEBHOOK_SECRET")
    verified = verify_signature(body, x_razorpay_signature, secret)
    if secret and not verified:
        raise HTTPException(401, "bad signature")
    try:
        payload = json.loads(body.decode("utf-8"))
    except (ValueError, UnicodeDecodeError) as exc:
        raise HTTPException(400, "invalid JSON") from exc
    event_id = event_identity(payload, x_razorpay_event_id)
    is_new, dupes = s.record_webhook(
        event_id, str(payload.get("event", "unknown")), payload, verified=verified
    )
    return {
        "event_id": event_id,
        "status": "recorded" if is_new else "duplicate_ignored",
        "duplicates_ignored": dupes,
        "verified": verified,
    }


@app.get("/webhooks/stats")
def webhook_stats(s: Store = Depends(store)) -> dict[str, int]:
    return s.webhook_stats()
