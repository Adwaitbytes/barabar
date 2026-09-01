"""Bank statement readers/writers in each bank's real export layout.

Detection is by header signature. Amounts are parsed with ``parse_inr`` (never a
float). Narrations go through the deterministic grammar; rows the grammar cannot
parse keep ``narration=None`` so the LLM fallback can be offered *only* there.
"""

from __future__ import annotations

import csv
import io
from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from barabar.core.models import Bank, BankTxn
from barabar.core.money import format_inr, parse_inr
from barabar.core.narration import parse_narration


@dataclass(frozen=True)
class Layout:
    bank: Bank
    header: tuple[str, ...]
    date_fmt: str
    date_col: str
    value_date_col: str | None
    narration_col: str
    debit_col: str | None
    credit_col: str | None
    amount_col: str | None = None  # Kotak-style single amount + Dr/Cr flag
    drcr_col: str | None = None
    balance_col: str | None = None
    ref_col: str | None = None


LAYOUTS: dict[Bank, Layout] = {
    Bank.HDFC: Layout(
        Bank.HDFC,
        (
            "Date",
            "Narration",
            "Chq./Ref.No.",
            "Value Dt",
            "Withdrawal Amt.",
            "Deposit Amt.",
            "Closing Balance",
        ),
        "%d/%m/%y",
        "Date",
        "Value Dt",
        "Narration",
        "Withdrawal Amt.",
        "Deposit Amt.",
        balance_col="Closing Balance",
        ref_col="Chq./Ref.No.",
    ),
    Bank.ICICI: Layout(
        Bank.ICICI,
        (
            "S No.",
            "Value Date",
            "Transaction Date",
            "Cheque Number",
            "Transaction Remarks",
            "Withdrawal Amount (INR )",
            "Deposit Amount (INR )",
            "Balance (INR )",
        ),
        "%d/%m/%Y",
        "Transaction Date",
        "Value Date",
        "Transaction Remarks",
        "Withdrawal Amount (INR )",
        "Deposit Amount (INR )",
        balance_col="Balance (INR )",
        ref_col="Cheque Number",
    ),
    Bank.SBI: Layout(
        Bank.SBI,
        (
            "Txn Date",
            "Value Date",
            "Description",
            "Ref No./Cheque No.",
            "Debit",
            "Credit",
            "Balance",
        ),
        "%d %b %Y",
        "Txn Date",
        "Value Date",
        "Description",
        "Debit",
        "Credit",
        balance_col="Balance",
        ref_col="Ref No./Cheque No.",
    ),
    Bank.AXIS: Layout(
        Bank.AXIS,
        ("Tran Date", "CHQNO", "PARTICULARS", "DR", "CR", "BAL", "SOL"),
        "%d-%m-%Y",
        "Tran Date",
        None,
        "PARTICULARS",
        "DR",
        "CR",
        balance_col="BAL",
        ref_col="CHQNO",
    ),
    Bank.KOTAK: Layout(
        Bank.KOTAK,
        ("Sl. No.", "Date", "Description", "Chq / Ref number", "Amount", "Dr / Cr", "Balance"),
        "%d/%m/%Y",
        "Date",
        None,
        "Description",
        None,
        None,
        amount_col="Amount",
        drcr_col="Dr / Cr",
        balance_col="Balance",
        ref_col="Chq / Ref number",
    ),
}


def detect_bank(header: Sequence[str]) -> Bank | None:
    cleaned = tuple(h.strip() for h in header)
    for bank, layout in LAYOUTS.items():
        if cleaned == layout.header:
            return bank
    for bank, layout in LAYOUTS.items():  # tolerate extra trailing columns / spacing
        if all(any(h.strip().lower() == want.lower() for h in header) for want in layout.header):
            return bank
    return None


def _money(cell: str | None) -> int:
    if cell is None or not cell.strip():
        return 0
    return parse_inr(cell)


def read_bank_rows(
    rows: Iterable[Sequence[str]], *, bank: Bank | None = None, source_file: str = "upload.csv"
) -> list[BankTxn]:
    it = iter(rows)
    header = [h for h in next(it)]
    detected = bank or detect_bank(header)
    if detected is None:
        raise ValueError(f"unknown bank statement layout: {header}")
    layout = LAYOUTS[detected]
    idx = {name.strip().lower(): i for i, name in enumerate(header)}

    def col(row: Sequence[str], name: str | None) -> str | None:
        if name is None:
            return None
        i = idx.get(name.lower())
        return row[i] if i is not None and i < len(row) else None

    out: list[BankTxn] = []
    for n, row in enumerate(it, start=2):
        if not any(c.strip() for c in row):
            continue
        posted = datetime.strptime(
            (col(row, layout.date_col) or "").strip(), layout.date_fmt
        ).date()
        vd_raw = col(row, layout.value_date_col)
        value = (
            datetime.strptime(vd_raw.strip(), layout.date_fmt).date()
            if vd_raw and vd_raw.strip()
            else posted
        )
        if layout.amount_col:
            amount = _money(col(row, layout.amount_col))
            flag = (col(row, layout.drcr_col) or "").strip().upper()
            credit, debit = (amount, 0) if flag.startswith("CR") else (0, amount)
        else:
            debit, credit = _money(col(row, layout.debit_col)), _money(col(row, layout.credit_col))
        narration = (col(row, layout.narration_col) or "").strip()
        bal_raw = col(row, layout.balance_col)
        out.append(
            BankTxn(
                bank_txn_id=f"bank_{n - 1:05d}",
                bank=detected,
                value_date=value,
                posted_date=posted,
                narration_raw=narration,
                narration=parse_narration(narration, detected),
                credit=credit,
                debit=debit,
                balance_after=_money(bal_raw) if bal_raw and bal_raw.strip() else None,
                source_file=source_file,
                row_no=n - 1,
            )
        )
    return out


def read_bank_csv(
    text: str, *, bank: Bank | None = None, source_file: str = "upload.csv"
) -> list[BankTxn]:
    return read_bank_rows(csv.reader(io.StringIO(text)), bank=bank, source_file=source_file)


def read_bank_file(path: Path, *, bank: Bank | None = None) -> list[BankTxn]:
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError(f"{path.name}: workbook has no active sheet")
        rows = [
            [
                "" if c is None else (c.strftime("%d/%m/%Y") if isinstance(c, datetime) else str(c))
                for c in r
            ]
            for r in ws.iter_rows(values_only=True)
        ]
        return read_bank_rows(rows, bank=bank, source_file=path.name)
    return read_bank_csv(path.read_text(encoding="utf-8-sig"), bank=bank, source_file=path.name)


def _fmt(d: date, fmt: str) -> str:
    return d.strftime(fmt)


def write_bank_csv(txns: Iterable[BankTxn], bank: Bank) -> str:
    layout = LAYOUTS[bank]
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\n")
    w.writerow(layout.header)
    for i, t in enumerate(txns, start=1):
        cells: dict[str, str] = {
            layout.date_col: _fmt(t.posted_date, layout.date_fmt),
            layout.narration_col: t.narration_raw,
        }
        if layout.value_date_col:
            cells[layout.value_date_col] = _fmt(t.value_date, layout.date_fmt)
        if layout.amount_col and layout.drcr_col:
            cells[layout.amount_col] = format_inr(t.credit or t.debit, symbol=False)
            cells[layout.drcr_col] = "CR" if t.credit else "DR"
        else:
            cells[layout.debit_col or ""] = format_inr(t.debit, symbol=False) if t.debit else ""
            cells[layout.credit_col or ""] = format_inr(t.credit, symbol=False) if t.credit else ""
        if layout.balance_col and t.balance_after is not None:
            cells[layout.balance_col] = format_inr(t.balance_after, symbol=False)
        if layout.ref_col:
            cells[layout.ref_col] = ""
        if "S No." in layout.header:
            cells["S No."] = str(i)
        if "Sl. No." in layout.header:
            cells["Sl. No."] = str(i)
        if "SOL" in layout.header:
            cells["SOL"] = "1234"
        w.writerow([cells.get(h, "") for h in layout.header])
    return buf.getvalue()
